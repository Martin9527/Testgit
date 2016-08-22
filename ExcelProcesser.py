# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import json
class ExcelProcesser(object):
	"""docstring for ExcelProcesser"""
	def __init__(self):
		super(ExcelProcesser, self).__init__()
		#self.arg = arg
		self.wb = None
		self.ws = None
		self.i = 1
		self.project = {}
		self.project_list = []
		self.start_row = None
	def read_excel(self,file_name = None):
		if not file_name:
			return
		wb = load_workbook(file_name)
		if wb:
			self.wb = wb
			self.ws = wb.get_sheet_by_name('Sheet1')
	def creat_excel(self,out_filename = 'D:\\out.xlsx',wb = None):
		wb = Workbook()
		ws = wb.active 
		i = 1
		ws.cell(column = 1,row = 1,value ='项目名称')
		ws.cell(column = 2,row = 1,value ='建设性质')
		ws.cell(column = 3,row = 1,value ='建设年限')
		ws.cell(column = 4,row = 1,value ='建设规模及内容')
		ws.cell(column = 5,row = 1,value ='2015省级预算内')
		for project in self.project_list:
			i +=1
			self.add_project(project,ws,i)
		wb.save(out_filename)
	def add_project(self,project,ws,i):
		ws.cell(column = 1,row = i,value = project['项目名称'])
		ws.cell(column = 2,row = i,value =project['建设性质'])
		ws.cell(column = 3,row = i,value =project['建设年限'])
		ss = ''.join([project['建设规模'],'；',project['建设内容']])
		ws.cell(column = 4,row = i,value =ss)
		ws.cell(column = 5,row = i,value =project['资金来源'])
	def process_excel(self,ws):
		if not ws:
			return
		cells = ws['C']
		value = ''
		for cell in cells:
			if cell.value!=None:
				value += cell.value.encode('utf-8')
				if not self.start_row:
					self.start_row = cell.row
			else:
				if self.start_row != None:
					self.project['项目名称'] = value
					#process other item
					self.getOtherItem()
					self.project_list.append(self.project)
					self.project = {}
					value = ''
					self.start_row = None
				else:
					continue
	

	def getOtherItem(self):
		self.constructProp()
		self.constructScale()
		self.constructTime()
		self.fundSources()
		self.constructContent()
		

	def projectDepartment(self,column = 'B'):
		pass
	def projectName(self,column = 'C'):
		col = self.ws[column]
		for cell in col:
			if cell.value != None:
				value += cell.value.encode('utf-8')
				if not self.start_row:
					self.start_row = cell.row
			else:
				next_row = cell.row
				break
	
	def constructProp(self,column = 'D'):
		value = ''
		i = self.start_row
		cell = self.ws['D%d'%i]
		while cell.value!=None:
			value += cell.value.encode('utf-8')
			i+=1
			cell = self.ws['D%d'%i]
		self.project['建设性质'] = value
	def constructScale(self,column = 'E'):
		value = ''
		i = self.start_row
		cell = self.ws['E%d'%i]
		while cell.value!=None:
			value += cell.value.encode('utf-8')
			i+=1
			cell = self.ws['E%d'%i]
		self.project['建设规模'] = value
	def constructTime(self,column = 'F'):
		value = ''
		i = self.start_row
		cell = self.ws['F%d'%i]
		while cell.value!=None:
			value += cell.value.encode('utf-8')
			i+=1
			cell = self.ws['F%d'%i]
		self.project['建设年限'] = value
	def fundSources(self,column_s = 'H',column_m = 'I'):
		value = ''
		fund_dict = {}
		i = self.start_row
		cell = self.ws['H%d'%i]
		while cell.value!=None:
			name = cell.value.encode('utf-8')
			number = self.ws['I%d'%i].value
			fund_dict[name] = number
			i+=1
			cell = self.ws['H%d'%i]
		self.project['资金来源'] = '6000'#fund_dict
	def constructContent(self,column = 'M'):
		value = ''
		i = self.start_row
		cell = self.ws['M%d'%i]
		while cell.value!=None:
			value += cell.value.encode('utf-8')
			i+=1
			cell = self.ws['M%d'%i]
		self.project['建设内容'] = value

	def dump(self,lst):
		with open('D:\\test.txt','wb+') as fp:
   			fp.write(json.dumps(lst, ensure_ascii=False))
    	
	
if __name__ == '__main__':
	processer = ExcelProcesser()
	processer.read_excel('1.xlsx')
	processer.process_excel(processer.ws)
	processer.dump(processer.project_list)
	processer.creat_excel()
	
	


		